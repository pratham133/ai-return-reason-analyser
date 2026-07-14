import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

raw = pd.read_csv("categorized_returns.csv")[
    ["product", "category", "price", "return_reason", "business_category"]
]

root_cause_fix = {
    "Not as Described (Listing Mismatch)": (
        "Listing photos/copy overstate material quality or use inconsistent lighting for color",
        "Reshoot product photography in neutral lighting; add fabric/material spec field to listing",
        "High",
    ),
    "Damaged / Incomplete on Arrival": (
        "Weak packaging and/or logistics handling for fragile or liquid items",
        "Upgrade packaging spec for fragile/liquid SKUs; add tamper/seal QC check before dispatch",
        "High",
    ),
    "Size / Fit Mismatch": (
        "Size charts are inconsistent or not brand/category-specific",
        "Standardize size charts per brand & add a fit-quiz / model height-weight reference on listing",
        "High",
    ),
    "Product Malfunction": (
        "Insufficient QC on electronics/appliance batches before shipping",
        "Add a functional pre-dispatch test (power-on/charge check) for electronics & appliances",
        "Medium",
    ),
    "Build Quality / Durability": (
        "Supplier-side quality issue (stitching, coating, elastic) not caught in QC",
        "Tighten incoming QC threshold with supplier; add durability spot-checks per batch",
        "Medium",
    ),
    "Wrong Item Sent": (
        "Warehouse picking/packing errors during order fulfillment",
        "Add barcode-scan verification step at packing before an order is sealed",
        "Low-Medium",
    ),
}
# Order by return volume (descending), matches category_summary.csv
categories_ordered = [
    "Not as Described (Listing Mismatch)",
    "Damaged / Incomplete on Arrival",
    "Size / Fit Mismatch",
    "Product Malfunction",
    "Build Quality / Durability",
    "Wrong Item Sent",
]

wb = Workbook()

# ---------- Sheet 1: Raw Data ----------
ws_raw = wb.active
ws_raw.title = "Raw Data"
headers = ["Product", "Category", "Price (Rs)", "Return Reason", "Business Category"]
ws_raw.append(headers)
for cell in ws_raw[1]:
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E7D32")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in raw.itertuples(index=False):
    ws_raw.append(list(row))

for col_idx, width in zip(range(1, 6), [22, 16, 12, 45, 30]):
    ws_raw.column_dimensions[get_column_letter(col_idx)].width = width
for r in range(2, ws_raw.max_row + 1):
    for c in range(1, 6):
        ws_raw.cell(row=r, column=c).font = Font(name="Arial", size=10)
        ws_raw.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_raw.freeze_panes = "A2"

n_rows = len(raw)
last_raw_row = n_rows + 1  # data starts row 2

# ---------- Sheet 2: Summary Dashboard ----------
ws = wb.create_sheet("Summary Dashboard")
ws["A1"] = "E-COMMERCE RETURN REASON ANALYSIS"
ws["A1"].font = Font(name="Arial", bold=True, size=16, color="2E7D32")
ws["A2"] = "Sample dataset: 50 return reasons across apparel, electronics, footwear, home & beauty categories"
ws["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

summary_headers = [
    "Business Category", "No. of Returns", "% of Total",
    "Avg. Order Value (Rs)", "Root Cause", "Recommended Fix", "Priority"
]
header_row = 4
for i, h in enumerate(summary_headers, start=1):
    cell = ws.cell(row=header_row, column=i, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E7D32")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, cat in enumerate(categories_ordered, start=1):
    r = header_row + i
    root_cause, fix, priority = root_cause_fix[cat]
    ws.cell(row=r, column=1, value=cat)
    # Live formulas -- recalculate if Raw Data changes
    ws.cell(row=r, column=2, value=(
        f'=COUNTIF(\'Raw Data\'!E$2:E${last_raw_row},A{r})'
    ))
    ws.cell(row=r, column=3, value=(
        f'=B{r}/SUM(B{header_row+1}:B{header_row+len(categories_ordered)})'
    ))
    ws.cell(row=r, column=3).number_format = "0.0%"
    ws.cell(row=r, column=4, value=(
        f'=AVERAGEIF(\'Raw Data\'!E$2:E${last_raw_row},A{r},\'Raw Data\'!C$2:C${last_raw_row})'
    ))
    ws.cell(row=r, column=4).number_format = '"Rs "#,##0'
    ws.cell(row=r, column=5, value=root_cause)
    ws.cell(row=r, column=6, value=fix)
    ws.cell(row=r, column=7, value=priority)
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    prio_cell = ws.cell(row=r, column=7)
    prio_fill = {"High": "FFCDD2", "Medium": "FFF9C4", "Low-Medium": "FFE0B2"}.get(priority, "FFFFFF")
    prio_cell.fill = PatternFill("solid", fgColor=prio_fill)
    prio_cell.alignment = Alignment(horizontal="center", vertical="center")

total_row = header_row + len(categories_ordered) + 1
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
ws.cell(row=total_row, column=2, value=f"=SUM(B{header_row+1}:B{total_row-1})").font = Font(name="Arial", bold=True)

col_widths = {"A": 30, "B": 14, "C": 12, "D": 18, "E": 38, "F": 42, "G": 12}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Note documenting the source of root cause / fix / priority (qualitative, analyst judgment)
note_row = total_row + 2
ws.cell(row=note_row, column=1,
        value="Note: 'No. of Returns', '% of Total' and 'Avg. Order Value' are live formulas pulling from the Raw Data sheet. "
              "Root Cause / Recommended Fix / Priority are analyst judgment based on reviewing each category's return reasons.")
ws.cell(row=note_row, column=1).font = Font(name="Arial", size=9, italic=True, color="666666")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

# ---------- Chart ----------
chart = BarChart()
chart.type = "bar"
chart.title = "Returns by Category"
chart.y_axis.title = "Category"
chart.x_axis.title = "No. of Returns"
data = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + len(categories_ordered))
cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(categories_ordered))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 18
ws.add_chart(chart, f"A{note_row + 3}")

wb.save("AI_Return_Reason_Analysis.xlsx")
print("saved")
